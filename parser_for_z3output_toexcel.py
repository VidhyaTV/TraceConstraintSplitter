import openpyxl
import os
import re
import sys

from openpyxl import Workbook

#open a workbook
wb=Workbook();
#row number to start with in the worksheet
row_number=1;
#open a worksheet for writing
#for ws in wb.worksheets:
ws=wb.worksheets[0];

#for each folder in the current directory (run folder-assuming you place this script it the run folder) that has 'batchlength' in its name
loc=os.getcwd();#get current working directory
batch_folders=os.listdir(loc);#list of all sub-entries in the current directory
for batch_folder in batch_folders:
#{
	batch_folder_path=os.path.join(loc,batch_folder);#path of the sub-entry
	if ((os.path.isdir(batch_folder_path)) and ('batchlength' in batch_folder)):#if subentry is a directory and has the term 'batchlength' in it
	#{
		print("In folder :"+batch_folder+"\n");
		#get batchlength value from the folder's name and write it to row_number-column 1
		m1=re.search('batchlength(\d+)',batch_folder);
		if m1:
		#{
			ws.cell(row=row_number,column=1).value="batchlength="+m1.group(1);
			#go to next row: increment row_number #mostly everytime after you write
			#skip next row: i.e. set row_number=row_number+1 #for excel-readability
			row_number=row_number+2;
			#write 'run1' in row_number,column=2
			ws.cell(row=row_number,column=2).value="run1";
			#write 'run2' in row_number,column=3
			ws.cell(row=row_number,column=3).value="run2";
			#write 'run3' in row_number,column=4
			ws.cell(row=row_number,column=4).value="run3";
			#go to next row: increment row_number #mostly everytime after you write
			row_number=row_number+1;
			#remember this new row number as variable 'anchor': anchor=row_number; so that you can use it later to start writing 'fromxtoy' batchlength ranges
			anchor=row_number;
			
			#for each file in this batchfolder that has the name 'z3output'#there is only one such file
			files=os.listdir(batch_folder_path);#all files in the current batchfolder
			for file in files:
			#{
				if ('z3output' in file):#only if file name has z3output in it:
				#{
					#set [(run|column)_counter] to 1 #this variable will be used to move rightwards i.e. across columns
					column_number=1;
					#parse file line by line
					file_path=os.path.join(batch_folder_path,file);
					with open(file_path) as inp_z3outputfile:
					#{
						for line in inp_z3outputfile:
						#{
							#if line contains ':time'
							if ':time' in line:
							#{
								m2= re.search('\s:time(\s+)(\d+).(\d+)\)',line);
								if m2:
								#{
									#extract number that occurs after multiple spaces and before the losing bracket
									time_value1=m2.group(2);#value before decimal pt
									time_value2=m2.group(3);#value after decimal pt
									time_value=float(time_value1+'.'+time_value2);
									
									#if [(run|column)_counter] < 4
									if column_number < 4:
										#write it to row_number, column [(run|column)_counter]+1
										ws.cell(row=row_number, column=column_number+1).value=time_value;#+1 to give space to write batch ranges in the first column
										#increment [(run|column)_counter]
										column_number=column_number+1;
									else:
										#go to next row: increment row_number
										row_number=row_number+1;
										#set [(run|column)_counter] to 1
										column_number=1;										
										#write it to row_number, column [(run|column)_counter]+1
										ws.cell(row=row_number, column=column_number+1).value=time_value;#+1 to give space to write batch ranges in the first column
										#increment [(run|column)_counter]
										column_number=column_number+1;
								#}
								else:
								#{
									print("line has time but no time value");
									sys.exit();
								#}
							#}
							#else if line contains '*from(d+)to(d+)*'
							elif 'from' in line:
							#{
								#extract the from and to values
								m3=re.search('from(\d+)',line);
								m4=re.search('_to(\d+)',line);
								if m3 and m4:
								#{
									fromtime=m3.group(1);
									totime=m4.group(1);
									#write 'fromtovalues' to row_number column 1
									ws.cell(row=anchor,column=1).value=fromtime+' to '+totime;
									#go to next row: increment row_number #mostly everytime after you write
									anchor=anchor+1;#move anchor down as you write, will be needed if 'line containing smt-filename:"running...smt"' is encountered again
								#}
								else:
								#{
									print("line has time but no time value");
									sys.exit();
								#}
							#}
						#}
					#}
				#}
			#}
		#}
	#}
	row_number=row_number+2;# for two empty rows before adding z3runtimes for next batchlength from z3output file in next batchlength-folder
#}
wb.save('test.xlsx');#save all additions/changes to the target file